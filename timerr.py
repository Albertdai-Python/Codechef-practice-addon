from openpyxl import load_workbook
from tkinter import *
from tkinter.ttk import *
import os
import time
from datetime import datetime
def get_maximum_rows(*, s):
    rows = 0
    for max_row, row in enumerate(s, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows
def st():
    global st
    st=int(time.time())
def sp():
    global sp
    sp=int(time.time())
    s=ent_1.get().upper()
    m=float(ent_2.get())
    t=float(ent_3.get())
    r=int(ent_4.get())
    a=get_maximum_rows(s=ws)+1
    print(a)
    ws[f'A{a}']=s
    ws.cell(a,1).hyperlink=f'https://www.codechef.com/submit/{s}'
    ws[f'B{a}']=r
    ws[f'C{a}']=str(result.get())
    ws[f'D{a}']=m
    ws[f'E{a}']=t
    ws[f'F{a}']=sp-st
    wb.save("data.xlsx")
    ent_1.delete(0, END)
    ent_2.delete(0, END)
    ent_3.delete(0, END)
    ent_4.delete(0, END)
    
w=Tk()
result=StringVar()
l1=Label(w,text="Problem Name").pack()
ent_1=Entry(w)
ent_1.pack()
l4=Label(w,text="Rating").pack()
ent_4=Entry(w)
ent_4.pack()
l2=Label(w,text="Memory").pack()
ent_2=Entry(w)
ent_2.pack()
l3=Label(w,text="Time").pack()
ent_3=Entry(w)
ent_3.pack()
l4=Label(w,text="Language").pack()
cb=Combobox(w,textvariable=result)
cb['values']=["Python3","Pypy3","C++17","C++14","Java"]
cb.pack()
start=Button(w,text="Start",command=st)
start.pack()
end=Button(w,text="End",command=sp)
end.pack()
wb=load_workbook("data.xlsx")
ws=wb.worksheets[0]
st()
